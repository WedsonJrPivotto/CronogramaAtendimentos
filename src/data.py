from openpyxl import load_workbook
from openpyxl.styles import Font
import shutil
import os

class CredentialsError(Exception):
  """Exceção personalizada para erros de credenciais."""
  def __init__(self, mensagem="Erro nas credenciais fornecidas."):
    self.mensagem = mensagem
    super().__init__(self.mensagem)

class Planilha:
  @staticmethod
  def cnx_data():
    caminho_original = './data/Atendimentos.xlsx'
    caminho_backup = './data/backup/Atendimentos.xlsx'
    try:
      wb = load_workbook(filename=caminho_original)
    except FileNotFoundError:
      if os.path.exists(caminho_backup):
        shutil.copy(caminho_backup, caminho_original)
        wb = load_workbook(filename=caminho_original)
    return wb

  @staticmethod
  def credenciais_data(nome:str, mes:str, novo:bool = False):
    if not nome == '':
      wb = Planilha.cnx_data()
      sheet = wb['Planilha1']
      sheet['A1'] = f"Atendente: {nome}"
      sheet['D1'] = f"Mês: {mes}"

      fonte = Font(bold=True, underline='single') 
      sheet['A1'].font = fonte
      sheet['D1'].font = fonte

      wb.save('./data/Atendimentos.xlsx')
      if novo:
        wb_bk = load_workbook('./data/backup/Atendimentos.xlsx')
        sheet_usuarios = wb_bk['Planilha2']
        linha = 1
        while sheet_usuarios[f'A{linha}'].value is not None:
          linha += 1
        sheet_usuarios[f'A{linha}'] = nome
        wb_bk.save('./data/backup/Atendimentos.xlsx')
    else:
      raise CredentialsError()
    
    

  @staticmethod
  def add_atendimento_data(cliente:str, date:str, hr_ini:str, hr_fin:str, desc:str):
    wb = Planilha.cnx_data()
    sheet = wb['Planilha1']
    linha = 3
    while sheet[f'A{linha}'].value is not None or sheet[f'B{linha}'].value is not None:
      linha += 1

    # Adicionar os valores na linha encontrada
    sheet[f'A{linha}'] = cliente
    sheet[f'B{linha}'] = date
    sheet[f'C{linha}'] = hr_ini
    sheet[f'D{linha}'] = hr_fin
    sheet[f'E{linha}'] = desc

    wb.save('./data/Atendimentos.xlsx')

  @staticmethod
  def buscar_usuários():
    caminho_backup = './data/backup/Atendimentos.xlsx'
    wb = load_workbook(filename=caminho_backup)
    lista_usuarios = ['']
    sheet = wb['Planilha2']
    linha = 1
    while sheet[f'A{linha}'].value is not None:
      lista_usuarios.append(str(sheet[f'A{linha}'].value))
      linha += 1
    return lista_usuarios